from openpyxl import Workbook, load_workbook

PATH='C:/Users/User/Downloads/WHO-COVID-19-global-data.xlsx'
wb = load_workbook(PATH)
ws = wb.active
i=2
while i<260000:
    if ws['B'+str(i)].value==ws['B'+str(i+1)].value:
        i+=1
    else:
        country = ws['B'+str(i)].value
        fcase = ws['F'+str(i)].value
        hdeath = ws['H'+str(i)].value
        if fcase == 0 :
            d =0
        else:
            d = hdeath/fcase*100
            d = round(d , 2)
        print('    ',country ,': { cases: ',fcase,', death: ', hdeath,', density: ',d,' },',sep='' )
        i+=1
        # print(i)
print('結束')
# print(ws['B'+str(i)].value)